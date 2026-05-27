"""Export service for CSV exports and Telegram backups."""

import csv
import io
from datetime import datetime
from typing import List, Dict, Optional
from decimal import Decimal

import httpx
from sqlalchemy.orm import Session

from app.models.trade import Trade
from app.core.config import settings


class ExportService:
    """Handle trade exports to CSV and Telegram backups."""

    def __init__(self, db: Session):
        self.db = db

    def _format_decimal(self, value: Optional[Decimal]) -> str:
        """Format Decimal for CSV export."""
        if value is None:
            return ""
        return str(value)

    def _format_datetime(self, value: Optional[datetime]) -> str:
        """Format datetime for CSV export."""
        if value is None:
            return ""
        return value.strftime("%Y-%m-%d %H:%M:%S")

    def _trade_to_dict(self, trade: Trade) -> Dict[str, str]:
        """Convert Trade model to CSV-compatible dict."""
        return {
            "symbol": trade.symbol,
            "direction": trade.direction,
            "entry_price": self._format_decimal(trade.entry_price),
            "quantity": self._format_decimal(trade.quantity),
            "entry_time": self._format_datetime(trade.entry_time),
            "exit_price": self._format_decimal(trade.exit_price),
            "exit_time": self._format_datetime(trade.exit_time),
            "fees": self._format_decimal(trade.fees),
            "setup": trade.setup or "",
            "tactic": trade.tactic or "",
            "stop_price": self._format_decimal(trade.stop_price),
            "target_price": self._format_decimal(trade.target_price),
            "r_multiple": self._format_decimal(trade.r_multiple),
            "status": trade.status or "",
            "notes": trade.notes or "",
            "pnl": self._format_decimal(trade.pnl),
            "exit_reason": trade.exit_reason or "",
            "review_notes": trade.review_notes or "",
            "tags": trade.tags or "",
            "review_tags": trade.review_tags or "",
        }

    def export_trades_to_csv(
        self,
        from_date: Optional[str] = None,
        to_date: Optional[str] = None,
        status: Optional[str] = None,
        user_id: Optional[int] = None,
    ) -> str:
        """Export trades to CSV string.
        
        Args:
            from_date: ISO date string (YYYY-MM-DD) for start date filter
            to_date: ISO date string (YYYY-MM-DD) for end date filter  
            status: Filter by trade status
            
        Returns:
            CSV content as string
        """
        query = self.db.query(Trade)
        if user_id is not None:
            query = query.filter(Trade.user_id == user_id)

        # Apply filters
        if status:
            query = query.filter(Trade.status == status)
        
        if from_date:
            from_datetime = datetime.strptime(from_date, "%Y-%m-%d")
            query = query.filter(Trade.entry_time >= from_datetime)
        
        if to_date:
            to_datetime = datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            query = query.filter(Trade.entry_time <= to_datetime)
        
        # Exclude deleted trades by default
        query = query.filter(Trade.status != "deleted")
        
        trades = query.order_by(Trade.entry_time.asc()).all()
        
        if not trades:
            return ""
        
        # Write to CSV
        output = io.StringIO()
        fieldnames = [
            "symbol", "direction", "entry_price", "quantity", "entry_time",
            "exit_price", "exit_time", "fees", "setup", "tactic",
            "stop_price", "target_price", "r_multiple", "status", "notes",
            "pnl", "exit_reason", "review_notes", "tags", "review_tags"
        ]
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for trade in trades:
            writer.writerow(self._trade_to_dict(trade))
        
        return output.getvalue()

    def export_trades_to_xlsx(
        self,
        from_date: Optional[str] = None,
        to_date: Optional[str] = None,
        status: Optional[str] = None,
        user_id: Optional[int] = None,
    ) -> bytes:
        """Export trades to XLSX bytes."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        query = self.db.query(Trade)
        if user_id is not None:
            query = query.filter(Trade.user_id == user_id)
        if status:
            query = query.filter(Trade.status == status)
        if from_date:
            query = query.filter(Trade.entry_time >= datetime.strptime(from_date, "%Y-%m-%d"))
        if to_date:
            query = query.filter(Trade.entry_time <= datetime.strptime(to_date, "%Y-%m-%d"))
        query = query.filter(Trade.status != "deleted")
        trades = query.order_by(Trade.entry_time.asc()).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trades"

        headers = [
            "Symbol", "Entry Price", "Quantity", "Entry Time", "Exit Price",
            "Exit Time", "Fees", "PnL", "Setup", "Tactic", "Stop Price",
            "Target Price", "R-Multiple", "Status", "Notes",
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C97A3F", end_color="C97A3F", fill_type="solid")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r, trade in enumerate(trades, 2):
            ws.cell(row=r, column=1, value=trade.symbol)
            ws.cell(row=r, column=2, value=float(trade.entry_price) if trade.entry_price else 0)
            ws.cell(row=r, column=3, value=float(trade.quantity) if trade.quantity else 0)
            ws.cell(row=r, column=4, value=trade.entry_time.strftime("%Y-%m-%d %H:%M") if trade.entry_time else "")
            ws.cell(row=r, column=5, value=float(trade.exit_price) if trade.exit_price else "")
            ws.cell(row=r, column=6, value=trade.exit_time.strftime("%Y-%m-%d %H:%M") if trade.exit_time else "")
            ws.cell(row=r, column=7, value=float(trade.fees) if trade.fees else 0)
            ws.cell(row=r, column=8, value=float(trade.pnl) if trade.pnl else "")
            ws.cell(row=r, column=9, value=trade.setup or "")
            ws.cell(row=r, column=10, value=trade.tactic or "")
            ws.cell(row=r, column=11, value=float(trade.stop_price) if trade.stop_price else "")
            ws.cell(row=r, column=12, value=float(trade.target_price) if trade.target_price else "")
            ws.cell(row=r, column=13, value=float(trade.r_multiple) if trade.r_multiple else "")
            ws.cell(row=r, column=14, value=trade.status or "")
            ws.cell(row=r, column=15, value=trade.notes or "")

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 30)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def send_telegram_backup(
        self,
        csv_content: str,
        chat_id: str,
        bot_token: str,
        summary_text: str = "Daily Trading Journal Backup"
    ) -> bool:
        """Send CSV backup to Telegram.
        
        Args:
            csv_content: CSV string to send
            chat_id: Telegram chat ID
            bot_token: Telegram bot token
            summary_text: Optional summary text
            
        Returns:
            True if successful, False otherwise
        """
        if not csv_content:
            return False
            
        # Create a binary file-like object from the CSV content
        files = {
            "document": ("trading_journal_backup.csv", io.BytesIO(csv_content.encode("utf-8")), "text/csv")
        }
        
        data = {
            "chat_id": chat_id,
            "caption": summary_text
        }
        
        url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
        
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                response = await client.post(url, files=files, data=data)
                response.raise_for_status()
                return True
        except Exception:
            return False